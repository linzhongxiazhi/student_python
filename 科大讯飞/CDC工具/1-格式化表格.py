import openpyxl as vb
from openpyxl.styles import Font, Alignment, Border, PatternFill
import time


def format_excel(url):
    """对排完序的表格进行格式优化"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    cell_title = sheet["1"]
    """字体"""
    font_title = Font(name="黑体", size=11, bold=True, color="000000")
    """对齐方式"""
    aaligment = Alignment(horizontal="center", vertical="center")
    """边框"""
    border = Border(left=None, right=None, top=None, bottom=None)
    """格式化首行"""
    for i in cell_title:
        i.font = font_title
        i.alignment = aaligment
        i.border = border
    """设置单元格的宽高"""
    sheet.row_dimensions[1].height = 20
    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 9
    sheet.column_dimensions["C"].width = 9
    sheet.column_dimensions["D"].width = 11
    sheet.column_dimensions["E"].width = 8
    sheet.column_dimensions["F"].width = 8
    sheet.column_dimensions["G"].width = 15
    sheet.column_dimensions["H"].width = 10
    sheet.column_dimensions["I"].width = 17
    sheet.column_dimensions["J"].width = 15
    sheet.column_dimensions["K"].width = 11
    sheet.column_dimensions["L"].width = 12
    excel.save(url)
    return ""


def inserter_column(url):
    """插入所需要的列"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    """插入所需的首行"""
    sheet.insert_cols(idx=1, amount=3)
    sheet.insert_cols(idx=10, amount=1)
    """给首行写入内容"""
    sheet["A1"].value = "事件级别"
    sheet["B1"].value = "医生"
    sheet["C1"].value = "卫生室"
    sheet["J1"].value = "医生对应病例数"
    excel.save(url)
    return ""


def mark_line(url,symptom):
    """查找需要标记出高亮的行并标出高亮"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    pattern_fill = PatternFill(fill_type="solid", fgColor="99ccff")
    """特殊情况"""
    pattern_fill_2 = PatternFill(fill_type="solid", fgColor="61ac85")
    symptom_2 = symptom.copy()
    special_symtom = ["发热","咳嗽"]
    for i in special_symtom:
        if i in symptom_2:
            symptom_2.remove(i)

    for i in sheet.iter_rows(min_row=2, max_row=101, min_col=9, max_col=9):
        for cell in i:
            if cell.value in symptom_2:
                for j in sheet[str(cell.row)]:
                    j.fill = pattern_fill
            if cell.value in special_symtom:
                for j in sheet[str(cell.row)]:
                    j.fill = pattern_fill_2
    excel.save(url)
    return ""


def marque(url,symptom):
    """标记高低"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    # """特殊情况处理"""
    symptom_2 = symptom.copy()
    special_symtom = ["发热","咳嗽"]
    for i in special_symtom:
        if i in symptom_2:
            symptom_2.remove(i)

    for i in sheet.iter_rows(min_row=2, max_row=101, min_col=9, max_col=11):
        for cell in i:
            maladies = sheet["I" + str(cell.row)].value  # maladies：症状
            cas_nomber = sheet["K" + str(cell.row)].value  # cas_nomber :病例数
            if cas_nomber == None:
                continue
            elif maladies == None:
                continue
            elif maladies in symptom_2:
                if cas_nomber < 3:
                    sheet["A" + str(cell.row)].value = 1
                elif cas_nomber >= 6:
                    sheet["A"+str(cell.row)].value = 3
            elif maladies in special_symtom:
                if  cas_nomber < 10:
                    sheet["A"+str(cell.row)].value = 1
            elif maladies not in symptom:
                if cas_nomber < 10:
                    sheet["A" + str(cell.row)].value = 1
                    continue
    excel.save(url)
    return ""


def inspecter_sheet(url):
    """检查工作表是否已经是格式化后的"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    if sheet["A1"].value == "事件级别":
        if sheet["B1"].value == "医生":
            if sheet["C1"].value == "卫生室":
                if sheet["J1"].value == "医生对应病例数":
                    sheet.delete_cols(1)
                    sheet.delete_cols(1)
                    sheet.delete_cols(1)
                    sheet.delete_cols(7)
    excel.save(url)
    return ""
"""<--------------------------------特殊情况-------------------------------->"""

def special(url):
    """特殊地区筛选"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    symptom_1 = "呼吸道症候群"  # symptom:症状
    region = ["阜阳市","宿州市"]  # region ：地区
    pattern_fill = PatternFill(fill_type="solid", fgColor="ffffff")
    for i in sheet.iter_rows(min_row=2, max_row=101, min_col=9, max_col=9):
        for cell in i:
            maladies = sheet["I" + str(cell.row)].value  # maladies：症状
            cas_nomber = sheet["K" + str(cell.row)].value  # cas_nomber :病例数
            if maladies == symptom_1:
                if str(sheet["E" + str(cell.row)].value) not in region:
                    for j in sheet[str(cell.row)]:
                        j.fill = pattern_fill
                    if int(cas_nomber) < 10:
                        sheet["A" + str(cell.row)].value = 1
                    else:
                        sheet["A" + str(cell.row)].value = None
                elif str(sheet["E" + str(cell.row)].value) in region:
                    if 3 <= cas_nomber < 5:
                        sheet["A" + str(cell.row)].value = 2
                    elif cas_nomber < 3:
                        sheet["A" + str(cell.row)].value = 1
                    elif cas_nomber >= 5:
                        sheet["A" + str(cell.row)].value = 3
    excel.save(url)
    return ""
if __name__ == '__main__':
    symptom = ["呕吐", "腹痛", "腹泻", "呕吐伴腹泻", "腹泻血便", "腹泻水样便", "高热",  "眼红",  "急性黄疸","皮疹","发热伴胸闷","咳嗽","发热","出疹"]
    print("<---开始处理--->")
    url = input("请输入文件名的绝对路径：")
    time1 = time.time()
    inspecter_sheet(url)  # 检查工作表
    print("<---工作表已检查完毕--->")
    inserter_column(url)  # 插入所需的行
    print("<---所需行已插入--->")
    format_excel(url)  # 格式化首行
    print("<---首行已格式化完毕--->")
    mark_line(url,symptom)  # 标出高亮
    print("<---重点监控症状高亮已标注完毕--->")
    marque(url,symptom)  # 给不需要人工看的病例标出123
    # special(url)
    # print("<---已对特殊地区进行处理--->")
    print("<---工作表总体优化完毕--->")
    time2 = time.time()
    print("用时：{:.2f} 秒".format(time2-time1))