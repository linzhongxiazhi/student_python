import pandas as pd
import openpyxl as vb
import time
import re
import os
from openpyxl.styles import Font,Alignment,Side,Border,PatternFill,GradientFill

def changer(url):  # changer:转换
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    for i in sheet.iter_rows(min_row=2, max_row=101, max_col=1):
        for cell in i:
            if cell.value == None:
                continue
            elif cell.value == 1:
                cell.value = "低"
            elif cell.value == 2:
                cell.value = "中"
            elif cell.value == 3:
                cell.value = "高"
    excel.save(url)
    return ""

def get_url(hopital,cas,path = "./每日病例"):
    """获取所需要文件的路径"""
    hopital = ".*" + hopital + ".*\.xlsx"    # hopital :医院

    """查找目录下的所有文件"""
    files = os.listdir(path)

    """找到类似卫生院的所有文件路径"""
    hopital_confus_path = []  # confus:模糊
    for i in files:
        ret = re.findall(hopital, i)
        for i in ret:
            hopital_confus_path.append(path + "\\" + i)
    if len(hopital_confus_path) == 0:
        return print("查找不到文件路径，请检查您的输入是否有误！")

    """确定卫生院路径"""
    final_path = ""  # final:最终
    for i in hopital_confus_path:
        excel = vb.load_workbook(i)
        sheet = excel["Sheet0"]
        if sheet["E2"].value == cas:
            final_path = i
    return final_path

def screen_time(path,time):
    """筛选病例的时间"""
    """将excel表格转化为DataFrame表"""
    data_text = pd.read_excel(path)
    df = pd.DataFrame(data_text)
    """排序和筛选表格"""
    df_1 = df.sort_values(["就诊时间"], ascending=[True])
    df_2 = df[df['就诊时间'].str.contains(time)]
    """写入表格"""
    writer = pd.ExcelWriter(path)
    df_1.to_excel(writer, sheet_name='Sheet0', index=False)
    df_2.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    return ""

def screen_cas(url,symptom):
    """筛选病例"""
    excel = vb.load_workbook(url)
    sheet = excel["Sheet1"]
    sheet_2 = excel["Sheet0"]
    if sheet["B2"].value == None:
        abc={"" :{"":""}}
        return abc
    else:
        docteur = []  # docteur:医生
        hopita = []  # hopita：医院
        """将医生和卫生院存入列表"""
        for i in sheet.iter_rows(min_row=2, min_col=7, max_col=8):
            for cell in i:
                if cell.column == 7:
                    docteur.append(cell.value)
                if cell.column == 8:
                    hopita.append(cell.value)
        """开始分组"""
        dic_wsy = pd.DataFrame({
            "卫生院": docteur,
            "医生和病例数": hopita
        })
        Hopita_docteur = dic_wsy["卫生院"].groupby(dic_wsy["医生和病例数"]).groups
        """将下标转化为医生"""
        for k, v in Hopita_docteur.items():
            x = 0  # 下标计数器
            ls = []  # 临时存放医生
            for i in v:
                ls.append(docteur[i])
            Hopita_docteur[k] = ls
        """对医生进行计数"""
        for k, v in Hopita_docteur.items():
            a = {}  # 病例计数器
            for i in v:
                a[i] = v.count(i)
            Hopita_docteur[k] = a
        """对病例数没到5的医生进行筛选"""
        hopita_docteur_malade = {}  # malade：病人
        for k, v in Hopita_docteur.items():
            docteur_malade = {}
            for i in v:
                if v[i] >= 5:
                    docteur_malade[i] = v[i]
            hopita_docteur_malade[k] = docteur_malade
        """删除值为空的卫生院"""
        hopita_docteur_malade_2 = {}
        for k, v in hopita_docteur_malade.items():
            if len(v) != 0:
                hopita_docteur_malade_2[k] = hopita_docteur_malade[k]

        """判断是不是空值，如果是判断是不是需要监控的症状，如果是返回病例数最多的一个医生。"""
        """找出每家医院病例最多的医生"""
        if len(hopita_docteur_malade_2) == 0:
            # if sheet_2["E2"].value in symptom:
            for k, v in Hopita_docteur.items():
                b = []
                c = []
                for x, y in v.items():
                    b.append(x)
                    c.append(y)
                subscript = c.index(max(c))
                b_c = {}
                b_c[b[subscript]] = c[subscript]
                Hopita_docteur[k] = b_c
            """将字典转化为列表"""
            for k, v in Hopita_docteur.items():
                ls = []
                for x, y in v.items():
                    ls.append(x)
                    ls.append(y)
                Hopita_docteur[k] = ls
            """找出该地方病人最多的医生"""
            a = []
            b = []
            c = []
            for k, v in Hopita_docteur.items():
                # print(k,v)
                a.append(k)
                b.append(v[0])
                c.append(v[1])
            subscript = c.index(max(c))  # subscript：下标
            Hopita_Docteur = {}
            hopita_docteur_malade_2 = {}
            Hopita_Docteur[b[subscript]] = c[subscript]
            hopita_docteur_malade_2[a[subscript]] = Hopita_Docteur

        return hopita_docteur_malade_2

def mark_line(url,symptom):
    """查找需要标记出高亮的行并标出高亮"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    pattern_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    # """特殊情况"""
    # pattern_fill_2 = PatternFill(fill_type="solid", fgColor="654321")
    # special_symtom = ["发热","咳嗽"]
    # for i in special_symtom:
    #     if i in symptom:
    #         symptom.remove(i)

    for i in sheet.iter_rows(min_row=2, max_row=101, min_col=9, max_col=9):
        for cell in i:
            if cell.value in symptom:
                if sheet["J" + str(cell.row)].value == None:
                    continue
                else:
                    j_str = sheet["J" + str(cell.row)].value
                    ls_j = j_str.split(",")
                    if len(ls_j) > 1:
                        y = cell.row
                        for z in sheet[y]:
                            z.fill = pattern_fill
                    elif len(ls_j) <= 1:
                        if int(ls_j[0]) >= 5:
                            y = cell.row
                            for z in sheet[y]:
                                z.fill = pattern_fill
                        else:
                            continue
            #     """特殊情况"""
            # elif cell.value in special_symtom:
            #     j_str = sheet["J" + str(cell.row)].value
            #     ls_j = j_str.split(",")
            #     if len(ls_j) > 1:
            #         y = cell.row
            #         for z in sheet[y]:
            #             z.fill = pattern_fill_2
            #     elif len(ls_j) <= 1:
            #         if int(ls_j[0]) >= 10:
            #             y = cell.row
            #             for z in sheet[y]:
            #                 z.fill = pattern_fill_2
                    else:
                        continue
    excel.save(url)
    return ""

def rinse_mark(url):
    """把之前标蓝的清洗掉"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    pattern_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for i in sheet.iter_rows(min_row=2, max_row=101):
        for cell in i:
            cell.fill =pattern_fill
    excel.save(url)
    return ""

"""<--------------------------------特殊情况-------------------------------->"""
def special(url):
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    symptom = ["呼吸道症候群"]  # symptom:症状
    region = ["阜阳市","宿州市"]  # region ：地区
    pattern_fill = PatternFill(fill_type="solid", fgColor="ffffff")
    for i in sheet.iter_rows(min_row=2, max_row=101, min_col=9, max_col=9):
        for cell in i:
            if sheet["I" + str(cell.row)].value in symptom:
                if sheet["E" + str(cell.row)].value not in region:
                    for j in sheet[str(cell.row)]:
                        j.fill = pattern_fill
    excel.save(url)
    return ""

if __name__ == '__main__':
    print("<---开始处理--->")
    symptom = ["呕吐", "腹痛", "腹泻", "呕吐伴腹泻", "腹泻血便", "腹泻水样便", "高热", "眼红", "急性黄疸", "皮疹", "发热伴胸闷", "咳嗽", "发热", "出疹"]
    url =input("请输入要处理的文件的绝对路径:")
    path = input("请输入病例的文件夹名的绝对路径：")
    time1 = time.time()
    """把123转化成高中低"""
    changer(url)
    print("<---转化完毕--->")
    """打开工作表"""
    excel = vb.load_workbook(url)
    sheet = excel["1-每日监控告警明细"]
    excel_time = str(sheet["D2"].value)
    """读入卫生院和症状"""
    hopital = ""
    malade =""
    for cell in sheet["A"]:
        if  sheet["A" + str(cell.row)].value =="事件级别":
            continue
        elif sheet["A" + str(cell.row)].value == "低":
            continue
        elif sheet["A" + str(cell.row)].value == "中":
            continue
        elif sheet["A" + str(cell.row)].value == None:
            continue
        elif sheet["A" + str(cell.row)].value == "高":
            """获取卫生院名称和症状"""
            hopital = sheet["G" + str(cell.row)].value
            malade = sheet["I" + str(cell.row)].value
            """找到对应病例对应的绝对路径"""
            file_path = get_url(hopital, malade,path)
            """筛选病例对应的时间"""
            screen_time(file_path, excel_time)
            """筛选出医生"""
            clinique_docteur = screen_cas(file_path,symptom)  # clinique: 医院 docteur：医生
            print(sheet["G" + str(cell.row)].value+"+"+sheet["I" + str(cell.row)].value)
            print(clinique_docteur)
            """开始写入"""
            docteur = ""  # docteur：医生
            clinique = ""  # clinique：医院
            malade = ""  # malade：病人
            for k,v in clinique_docteur.items():
                for x,y in v.items():
                    clinique = k + "," + clinique
                    docteur = x + "," + docteur
                    malade = str(y) + "," + malade
            sheet["B" + str(cell.row)].value = docteur[0:-1]
            sheet["C" + str(cell.row)].value = clinique[0:-1]
            sheet["J" + str(cell.row)].value = str(malade)[0:-1]
    excel.save(url)
    rinse_mark(url)
    mark_line(url,symptom)
    # special(url)
    """打印结果"""
    time2 = time.time()
    print("搞完了，搞完了！")
    print("记得备份哦！")
    print("用时：{:.2f} 秒".format(time2-time1))




