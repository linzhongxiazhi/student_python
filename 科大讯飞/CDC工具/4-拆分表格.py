import openpyxl as vb
from openpyxl.styles import Font,Alignment,Side,Border,PatternFill,GradientFill
import time
url = input("请输入文件名的绝对路径：")
time1= time.time()
excel =vb.load_workbook(url)
sheet = excel["Sheet1"]
for i in sheet.iter_rows(min_row=2, max_col=1,max_row=201):
    for cell in i:
        if cell.value == None:
            continue
        else:
            doctor= cell.value.split(",")
            hospital = sheet["B" +str(cell.row)].value.split(",")
            patient =sheet["I" +str(cell.row)].value.split(",")
            if len(doctor) != 1:
                sheet.insert_rows(idx=cell.row+1, amount=len(doctor)-1)
                for i in range(len(doctor)):
                    sheet["A"+str(cell.row+i)].value = doctor[i]
                    sheet["B" + str(cell.row + i)].value = hospital[i]
                    sheet["C" + str(cell.row + i)].value = sheet["C"+str(cell.row)].value
                    sheet["D" + str(cell.row + i)].value = sheet["D"+str(cell.row)].value
                    sheet["E" + str(cell.row + i)].value = sheet["E"+str(cell.row)].value
                    sheet["F" + str(cell.row + i)].value = sheet["F"+str(cell.row)].value
                    sheet["G" + str(cell.row + i)].value = sheet["G"+str(cell.row)].value
                    sheet["H" + str(cell.row + i)].value = sheet["H"+str(cell.row)].value
                    sheet["I" + str(cell.row + i)].value = patient[i]
                    sheet["J" + str(cell.row + i)].value = sheet["J"+str(cell.row)].value
                    sheet["K" + str(cell.row + i)].value = sheet["K" + str(cell.row)].value
                    sheet["L" + str(cell.row + i)].value = sheet["L" + str(cell.row)].value
                    sheet["M" + str(cell.row + i)].value = sheet["M" + str(cell.row)].value
                    sheet["N" + str(cell.row + i)].value = sheet["N" + str(cell.row)].value
                    sheet["O" + str(cell.row + i)].value = sheet["O" + str(cell.row)].value
                    sheet["P" + str(cell.row + i)].value = sheet["P" + str(cell.row)].value
                    sheet["Q" + str(cell.row + i)].value = sheet["Q" + str(cell.row)].value
                    sheet["R" + str(cell.row + i)].value = sheet["R" + str(cell.row)].value
                    sheet["S" + str(cell.row + i)].value = sheet["S" + str(cell.row)].value
                    sheet["T" + str(cell.row + i)].value = sheet["T" + str(cell.row)].value
                    sheet["U" + str(cell.row + i)].value = sheet["U" + str(cell.row)].value
                    sheet["V" + str(cell.row + i)].value = sheet["V" + str(cell.row)].value
                    sheet["W" + str(cell.row + i)].value = sheet["W" + str(cell.row)].value
                    sheet["X" + str(cell.row + i)].value = sheet["X" + str(cell.row)].value
                    sheet["Y" + str(cell.row + i)].value = sheet["Y" + str(cell.row)].value
"""填充颜色"""
pattern_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
"""边框"""
side = Side(style="thin",color="FF000000")
border = Border(left=side, right=side, top=side, bottom=side)
for i in sheet[1]:
    i.border =border
for i in sheet.iter_rows(min_row=2):
    for cell in i:
        cell.fill = pattern_fill
        cell.border = border
"""设置单元格的宽高"""
sheet.row_dimensions[1].height = 20
sheet.column_dimensions["A"].width = 10
sheet.column_dimensions["B"].width = 27
sheet.column_dimensions["C"].width = 11
sheet.column_dimensions["D"].width = 9
sheet.column_dimensions["E"].width = 9
sheet.column_dimensions["F"].width = 27
sheet.column_dimensions["G"].width = 9
sheet.column_dimensions["H"].width = 18
sheet.column_dimensions["I"].width = 15
sheet.column_dimensions["J"].width = 10
sheet.column_dimensions["K"].width = 11


excel.save(url)
time2 = time.time()
print("下班，下班！")
print("用时：{:.2f} 秒".format(time2 - time1))