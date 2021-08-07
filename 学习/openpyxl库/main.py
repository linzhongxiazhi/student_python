import openpyxl as vb
from openpyxl.styles import Font,Alignment,Side,Border,PatternFill,GradientFill
import pandas as pd

def sort_file(url):
    """对文件进行排序
    file : 需要排序的工作表
    """
    data_text = pd.read_excel(url)
    df = pd.DataFrame(data_text)
    df_1 = df.sort_values(["告警关注度","城市名称","区县名称"], ascending=[False,True,True])
    writer = pd.ExcelWriter(url)
    df_1.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    return ""
def format_excel(url):
    """对排完序的表格进行格式优化"""
    excel = vb.load_workbook(url)
    sheet = excel.active
    cell_title = sheet["1"]
    font_title = Font(name="黑体", size=11, bold=True, color="000000")
    aaligment = Alignment(horizontal="center", vertical="center")
    border = Border(left=None, right=None, top=None, bottom=None)
    """颜色"""
    pattern_fill = PatternFill(fill_type="solid", fgColor="99ccff")
    for i in cell_title:
        i.font = font_title
        i.alignment = aaligment
        i.border = border
        i.fill = pattern_fill
    sheet.row_dimensions[1].height = 20
    sheet.column_dimensions["A"].width = 9
    sheet.column_dimensions["B"].width = 9
    sheet.column_dimensions["C"].width = 9
    sheet.column_dimensions["D"].width = 11
    sheet.column_dimensions["E"].width = 11
    sheet.column_dimensions["F"].width = 11
    sheet.column_dimensions["G"].width = 40
    sheet.column_dimensions["H"].width = 10
    sheet.column_dimensions["I"].width = 17
    sheet.column_dimensions["J"].width = 15
    sheet.column_dimensions["K"].width = 11
    sheet.column_dimensions["L"].width = 12
    sheet.title = "1-每日监控告警明细"
    excel.save(url)
    return ""


def inserter_column(url):
    """插入所需要的列"""
    excel = vb.load_workbook(url)
    excel_b = excel.active
    excel_b.insert_cols(idx=1, amount=3)
    excel_b.insert_cols(idx=10, amount=1)
    excel_C = excel["Sheet1"]
    excel_cella1 = excel_C["A1"]
    excel_cella1.value = "事件级别"
    excel_cellb1 = excel_C["B1"]
    excel_cellb1.value = "医生"
    excel_cellc1 = excel_C["C1"]
    excel_cellc1.value = "卫生院"
    excel_cellj1 = excel_C["J1"]
    excel_cellj1.value = "医生当日病例数"
    excel.save(url)
    return ""

def mark_line(url):
    """查找需要标记出高亮的行并标出高亮"""
    excel = vb.load_workbook(url)
    sheet = excel.active
    cell_I = sheet["I"]
    symptom = ["呕吐", "腹痛", "腹泻", "呕吐伴腹泻", "腹泻血便", "腹泻水样便", "发热伴出诊", "高热", "发热伴皮疹", "发热伴呕吐", "眼红", "食物中毒", "急性黄疸",
               "急性出血性结膜炎", "皮疹"]
    i = 0
    pattern_fill = PatternFill(fill_type="solid", fgColor="99ccff")
    for cell in cell_I:
        i += 1
        if i >= 102:
            break
        if cell.value in symptom:
            cell.fill = pattern_fill

    excel.save(url)
    return ""


if __name__ == '__main__':
    print("初始化中")
    print("正在读取表格")
    url ="E:\project\python\openpyxl库\表格\每日监控告警明细_20210621083113.xlsx"
    print("读取完毕，准备对表格进行排序")
    sort_file(url) #1.把表格进行排序
    print("排序完毕")
    print("准备插入所需的行")
    inserter_column(url)
    print("插入完毕")
    print("调整单元格大小")
    format_excel(url)
    print("单元格大小已调整完毕")
    print("准备进行重点监控症状高亮标注")
    mark_line(url)
    print("标注完毕")
