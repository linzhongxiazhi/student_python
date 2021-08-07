import openpyxl as vb
from openpyxl.styles import Font,Alignment,Side,Border,PatternFill,GradientFill
url ="D:\林钟\项目\表格\每日监控告警明细_20210619084512.xlsx"
excel = vb.load_workbook(url)
sheet = excel.active
cell_I = sheet["I"]
symptom = ["呕吐","腹痛","腹泻","呕吐伴腹泻","腹泻血便","腹泻水样便","发热伴出诊","高热","发热伴皮疹","发热伴呕吐","眼红","食物中毒","急性黄疸","急性出血性结膜炎","皮疹"]
i=0
pattern_fill = PatternFill(fill_type="solid", fgColor="99ccff")
for cell in cell_I:
    i += 1
    if i>=102:
        break
    if cell.value in symptom:
        cell.fill = pattern_fill

excel.save(url)